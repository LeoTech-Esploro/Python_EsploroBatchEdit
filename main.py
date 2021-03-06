import openpyxl
import xmlschema
from yattag import Doc, indent

#load Excel file
filepath = "C:\\Users\\sean1\\Desktop\\Sean Costello\\School\\University of La Verne\\Senior Year\\Spring Semester\\CMPS 471- Internship\\Wilson Library\\Project4\\sampleExcel.xlsx"
wrkbk = openpyxl.load_workbook(filepath)
ws = wrkbk.active

#input XML version information
version_info = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
segment_type = 'segment_type="External"'

#lists needed
column_headers_list:list = []
researcher_row_list:list = []
temp_list:list = []

#iterate through each row in workbook
row_counter = 0
for row in ws.iter_rows(min_row = 1, max_row = ws.max_row, min_col = 1, max_col = ws.max_column):
    #check to make sure cell is not NULL
    for cell in row:
        if cell.value == None:
            temp_list.append("N/A")
        else:
            temp_list.append(cell.value)
    #check to make sure entire row is not NULL
    counter = 0
    for i in temp_list:
        if i == "N/A":
            counter += 1
    if counter == ws.max_column:
        temp_list = []
    else:
        #check to see if row contains column headers
        if row_counter == 0:
            #add row to column_headers_list
            column_headers_list = temp_list
            temp_list = []
            row_counter += 1
        else:
            #add row to researcher_row_list
            researcher_row_list.append(temp_list)
            temp_list = []
            row_counter += 1
            
#template for XML document
"""
<?xml version="1.0" encoding="UTF-8" standalone="yes">
<users>
    <user>
        <is_researcher>True/False</is_researcher>
        <user_identifiers>
            <user_identifier segment_type="External">
                <id_type>OTHER_ID_1</id_type>
                <value>someone@example.com</value>
                <status>ACTIVE</status>
            </user_identifier>
        </user_identifiers>
        <researcher>
            <photo_url>https://img.com/example</photo_url>
        </researcher>
    </user>
</users>
"""

#find positions of columns
is_researcher_position = 0
user_identifier_position = 0
status_position = 0
photo_url_position = 0
header_position = 0
for header in column_headers_list:
    if header == "is_researcher":
        is_researcher_position = header_position
    elif header == "user_identifier":
        user_identifier_position = header_position
    elif header == "status":
        status_position = header_position
    elif header == "photo_url":
        photo_url_position = header_position
    header_position += 1

#generate format for the XML document according to the above template
doc, tag, text = Doc().tagtext()
for row in researcher_row_list:
    with tag("users"):
        with tag("user"):
            with tag(column_headers_list[is_researcher_position]):
                text(row[is_researcher_position])
            with tag(column_headers_list[user_identifier_position]):
                with tag(column_headers_list[user_identifier_position], segment_type):
                    with tag("id_type"):
                        text("OTHER_ID_1")
                    with tag("value"):
                        text(row[user_identifier_position])
                    with tag("status"):
                        text(row[status_position])
            with tag("researcher"):
                with tag("photo_url"):
                    text(row[photo_url_position])

result = indent(
    doc.getvalue(),
    indentation = '    ',
    indent_text = False
    )

#generate XML document
with open("XML_Output.xml", "w") as files:
    files.write(version_info)
    files.write("\n")
    files.write(result)
    print("XML Conversion Complete")

#validate XML document against Esploro schema files
schema = xmlschema.XMLSchema("C:\\Users\\sean1\\Desktop\\Sean Costello\\School\\University of La Verne\\Senior Year\\Spring Semester\\CMPS 471- Internship\\Wilson Library\\Project4\\Validation\\rest_user.xsd")
schema.validate("C:\\Users\\sean1\\Desktop\\Sean Costello\\School\\University of La Verne\\Senior Year\\Spring Semester\\CMPS 471- Internship\\Wilson Library\\Project4\\XML_Output.xml")
print("Validation Successful")
